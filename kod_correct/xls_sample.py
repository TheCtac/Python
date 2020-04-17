from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import styles

wb = load_workbook('test1.xlsx')

ws = wb.active

# change worksheet name
#ws.title = "Test"

# create new worksheets using the Workbook.create_sheet() method
#ws1 = wb.create_sheet("Test")

# add row
#ws.insert_rows(7)

# move ranges of cells within a worksheet (translate formula)
#ws.move_range("D4:F10", rows=-1, cols=2, translate=True)
 
#ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
#ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
#ws.page_setup.fitToHeight = 0
#ws.page_setup.fitToWidth = 1

# append from array
#for r in data:
#    ws.append(r)

#  group column
#ws.column_dimensions.group('A','D', hidden=False)

# add print area
# ws.print_area = 'A1:F10'

font = styles.Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
font_b = styles.Font(name='Calibri',
            size=11,
            bold=True,
            color='FF000000')


thin = styles.Side(border_style="thin", color="000000")            
border = styles.Border(outline = thin, top = thin, left=thin, right=thin, bottom=thin, vertical=thin, horizontal=thin)

alignment = styles.Alignment(horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0)

ws.merge_cells('B1:D2')
# show error if insert value into first

# row - строка, column - поле

ws['C8'] = str(ws['C9'].value) + '!'
ws.cell(row=1, column=2, value='huy')

b1 = ws['B1']
b1.font = font_b
b1.alignment = alignment
# заливка
#b1.fill = styles.PatternFill("solid", fgColor="EE2222")
b1.border = border


b3 = ws['B3']
b3.border = border


wb.save('test2.xlsx')