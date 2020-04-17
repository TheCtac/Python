from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import styles
import tkinter as tk
from tkinter import filedialog

''' for coordinate into sheet
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
xy = coordinate_from_string('A4') # returns ('A',4)
col = column_index_from_string(xy[0]) # returns 1
row = xy[1] # return 4 
'''
root = tk.Tk()
root.withdraw()
file_ = filedialog.askopenfilename()

file_type = file_.split('.')[-1]
if file_type != 'xls' and file_type != 'xlsx':
    print('unsupported file type')
    exit()

wb = load_workbook(file_)

ws = wb.active
kod_field = ws['D']
#count_ = len(kod_field)

def char_to_num(text): 
    var_arr = ['A','a','B','В',  'А','а','Б','б','В','в','Г','г','Д','д','Е','е','Є','є']
    num_arr = [ 1,  1,  2,  2,    1,  1,  2,  2,  3,  3,  4,  4,  5,  5,  6,  6,  7,  7]
    for var_ in var_arr:
        index_ = var_arr.index(var_)
        text = text.replace(var_, str(num_arr[index_]))

    return text   

for row in kod_field:
    if row.row > 6:
        kod_ = row.value
        g = ws['D'+ str(row.row)]
        kod_ = char_to_num(kod_).strip() + '00000'
        g.value = kod_[0:13]

file_out = file_.replace('.' + file_type, '_out.' + file_type)

wb.save(file_out)
