from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import styles
import tkinter as tk
from tkinter import filedialog

''' for coordinate into sheet
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
xy = coordinate_from_string('A4') # returns ('A',4)
col = column_index_from_string(xy[0]) # returns 1
row = xy[1] # return 4 
'''
root = tk.Tk()
root.withdraw()
file_ = filedialog.askopenfilename()

file_type = file_.split('.')[-1]
if file_type != 'xls' and file_type != 'xlsx':
    print('unsupported file type')
    exit()

wb_in = load_workbook(file_)
ws_in = wb_in.active

wb_out = Workbook()
ws_out = wb_out.active

pokaz_ = ws_in['D']
len_ = len(pokaz_)

print('there are '+str(len_)+' rows in document')

for row in pokaz_:
    pok_ = row.value
    if pok_ != ' ':
        print(str(row.row)+'/'+str(len_))

        kod_ = ws_in['A'+str(row.row)].value
        numb_ = ws_in['B'+str(row.row)].value
        a = ws_out['A'+str(row.row)]
        b = ws_out['B'+str(row.row)]
        c = ws_out['C'+str(row.row)]
        
        a.value = kod_
        b.value = numb_
        c.value = pok_

file_out = file_.replace('.' + file_type, '_out.' + file_type)

wb_out.save(file_out)
